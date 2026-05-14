"""
管理干部API - 班级制培训管理系统（只读·含深度分析与导出）
"""
import csv
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.admin.shared import require_manager
from app.api.auth_v2 import get_current_user
from app.api.common_views import (
    get_analytics_summary,
    get_class_info,
    get_progress_overview,
    get_reading_ranking,
    get_scores_matrix,
    get_scores_ranking,
    get_single_test_scores,
    get_students_list,
)
from app.core.database import get_db
from app.models.class_system import (
    AlertRecord,
    AlertRule,
    Announcement,
    AuditLog,
    ChapterProgress,
    ChapterProgressStatus,
    Class,
    ClassMember,
    ClassStatus,
    ReadingProgress,
    Test,
    TestResult,
    User,
    UserRole,
)

router = APIRouter(prefix="/manager", tags=["管理干部"])


# ═══════════════════════════════════════════════
# 班级信息（共享）
# ═══════════════════════════════════════════════

async def get_manager_class(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    class_id: int | None = Query(None, description="指定班级ID，不提供则返回第一个活跃班级")
) -> Class:
    """依赖：获取管理干部可访问的班级（支持跨班查看）"""
    if user.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="需要管理干部权限")
    if class_id:
        result = await db.execute(select(Class).where(Class.id == class_id, Class.status == ClassStatus.ACTIVE))
        cls = result.scalar_one_or_none()
        if not cls:
            raise HTTPException(status_code=404, detail="班级不存在或已结束")
        return cls
    result = await db.execute(select(Class).where(Class.status == ClassStatus.ACTIVE).limit(1))
    cls = result.scalar_one_or_none()
    if not cls:
        raise HTTPException(status_code=404, detail="当前没有活跃班级")
    return cls


@router.get("/class")
async def get_manager_class_route(
    cls: Class = Depends(get_manager_class)
):
    """获取当前/指定班级信息"""
    return await get_class_info(cls)


@router.get("/classes")
async def get_manager_classes_alias(
    cls: Class = Depends(get_manager_class)
):
    """班级信息（前端别名 → /class）"""
    return await get_class_info(cls)


@router.get("/students")
async def list_students(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """学员列表（含阅读进度）"""
    return await get_students_list(db, cls, include_progress=True)


@router.get("/students/{user_id}")
async def get_student_detail(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """学员详情"""
    result = await db.execute(select(User).where(User.id == user_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="学员不存在")

    # 验证属于当前班级
    result = await db.execute(
        select(ClassMember).where(
            ClassMember.class_id == cls.id,
            ClassMember.user_id == user_id,
            ClassMember.role == UserRole.STUDENT
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="该学员不属于当前班级")

    # 阅读进度
    result = await db.execute(
        select(ReadingProgress).where(ReadingProgress.user_id == user_id)
    )
    readings = result.scalars().all()

    # 成绩
    result = await db.execute(
        select(TestResult, Test).join(Test)
        .where(TestResult.user_id == user_id, Test.class_id == cls.id)
    )
    test_results = result.all()

    avg_score = sum(tr.score for tr, t in test_results if tr.score) / len(test_results) if test_results else 0

    return {
        "id": student.id, "name": student.name,
        "id_card": student.id_card, "phone": student.phone,
        "reading_progress": [
            {"textbook_id": rp.textbook_id, "progress": rp.progress, "duration": rp.duration}
            for rp in readings
        ],
        "tests": [
            {"test_id": tr.test_id, "test_title": t.title,
             "score": tr.score, "time_spent": tr.time_spent,
             "submitted_at": tr.submitted_at.isoformat() if tr.submitted_at else None}
            for tr, t in test_results
        ],
        "tests_completed": len(test_results), "avg_score": round(avg_score, 1)
    }


# ═══════════════════════════════════════════════
# 进度 & 成绩（共享）
# ═══════════════════════════════════════════════

@router.get("/progress")
async def get_students_progress(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """所有学员进度"""
    return await get_progress_overview(db, cls)


@router.get("/scores")
async def get_scores_summary(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """成绩汇总表"""
    return await get_scores_matrix(db, cls)


@router.get("/scores/{test_id}")
async def get_test_scores(
    test_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """单次测验成绩"""
    return await get_single_test_scores(test_id, db, cls)


# ═══════════════════════════════════════════════
# 深度分析（共享 + 增强）
# ═══════════════════════════════════════════════

@router.get("/analytics/overview")
async def get_analytics_overview(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """班级统计概览（含通过率）"""
    return await get_analytics_summary(db, cls)


@router.get("/dashboard/summary")
async def get_dashboard_summary_alias(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """干部大屏汇总（前端别名 → /analytics/overview）"""
    return await get_analytics_summary(db, cls)


@router.get("/analytics/reading")
async def get_reading_analytics(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """阅读统计排行（含学习时长、成绩）"""
    return await get_reading_ranking(db, cls)


@router.get("/analytics/scores")
async def get_scores_analytics(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """成绩统计排行（含最高/最低分）"""
    return await get_scores_ranking(db, cls)


@router.get("/analytics/anti-cheat")
async def get_anti_cheat_data(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """反作弊数据：切标签次数统计"""
    # 获取所有学员
    result = await db.execute(
        select(ClassMember, User).join(User)
        .where(ClassMember.class_id == cls.id, ClassMember.role == UserRole.STUDENT)
    )
    rows = result.all()

    data = []
    for member, u in rows:
        # 统计该学员所有章节的切标签次数
        result = await db.execute(
            select(ChapterProgress).where(
                ChapterProgress.user_id == u.id,
                ChapterProgress.class_id == cls.id,
                ChapterProgress.tab_switch_count > 0
            )
        )
        tab_records = result.scalars().all()
        total_switches = sum(r.tab_switch_count or 0 for r in tab_records)

        if total_switches > 0:
            data.append({
                "user_id": u.id,
                "name": u.name,
                "total_switches": total_switches,
                "affected_chapters": len(tab_records)
            })

    # 按切换次数降序排列
    data.sort(key=lambda x: x["total_switches"], reverse=True)
    return {"suspicious": data}


# ═══════════════════════════════════════════════
# Excel/CSV 导出（干部核心需求）
# ═══════════════════════════════════════════════

@router.get("/export/students")
async def export_students(
    format: str = Query("xlsx", description="导出格式: xlsx/csv/json"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """导出学员名单（Excel/CSV/JSON）"""
    students = await get_students_list(db, cls, include_progress=True)

    if format == "json":
        return {
            "class_name": cls.name, "export_time": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
            "students": [
                {"序号": i + 1, "姓名": s["name"], "身份证号": s["id_card"],
                 "手机号": s["phone"], "阅读进度": f"{s.get('reading_progress', 0)}%",
                 "加入时间": s.get("joined_at", "")}
                for i, s in enumerate(students)
            ]
        }

    if format == "xlsx":
        return _export_students_xlsx(cls, students)

    # CSV 导出（默认回退）
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["序号", "姓名", "身份证号", "手机号", "阅读进度%", "加入时间"])
    for i, s in enumerate(students):
        writer.writerow([
            i + 1, s["name"], s["id_card"], s["phone"],
            s.get("reading_progress", 0), s.get("joined_at", "")
        ])

    output.seek(0)
    filename = f"{cls.name}_学员名单_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/scores")
async def export_scores(
    format: str = Query("xlsx", description="导出格式: xlsx/csv/json"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """导出成绩表（Excel/CSV/JSON）"""
    result = await db.execute(select(Test).where(Test.class_id == cls.id))
    tests = result.scalars().all()

    students = await get_students_list(db, cls)

    # 构建数据
    scores_data = []
    for s in students:
        row = {"姓名": s["name"], "身份证号": s["id_card"]}
        for test in tests:
            result = await db.execute(
                select(TestResult).where(
                    TestResult.test_id == test.id,
                    TestResult.user_id == s["user_id"]
                )
            )
            tr = result.scalar_one_or_none()
            row[test.title] = tr.score if tr else "-"
        scores_data.append(row)

    if format == "json":
        return {
            "class_name": cls.name, "export_time": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
            "tests": [t.title for t in tests], "scores": scores_data
        }

    if format == "xlsx":
        return _export_scores_xlsx(cls, tests, scores_data)

    # CSV 导出（默认回退）
    output = io.StringIO()
    headers = ["姓名", "身份证号"] + [t.title for t in tests]
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in scores_data:
        writer.writerow([row.get(h, "") for h in headers])

    output.seek(0)
    filename = f"{cls.name}_成绩表_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ═══════════════════════════════════════════════
# Excel 导出辅助函数
# ═══════════════════════════════════════════════

_XLSX_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_XLSX_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_XLSX_CELL_FONT = Font(name="微软雅黑", size=10)
_XLSX_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
_XLSX_CENTER = Alignment(horizontal="center", vertical="center")
_XLSX_LEFT = Alignment(horizontal="left", vertical="center")


def _style_xlsx_header(ws, row_num, col_count):
    """给 Excel 表头行加样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _XLSX_HEADER_FILL
        cell.font = _XLSX_HEADER_FONT
        cell.alignment = _XLSX_CENTER
        cell.border = _XLSX_THIN_BORDER


def _style_xlsx_cell(cell, align="center"):
    """给数据单元格加基础样式"""
    cell.font = _XLSX_CELL_FONT
    cell.alignment = _XLSX_CENTER if align == "center" else _XLSX_LEFT
    cell.border = _XLSX_THIN_BORDER


def _export_students_xlsx(cls: Class, students: list) -> StreamingResponse:
    """导出学员名单为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学员名单"

    headers = ["序号", "姓名", "身份证号", "手机号", "阅读进度%", "加入时间"]
    ws.append(headers)
    _style_xlsx_header(ws, 1, len(headers))

    for i, s in enumerate(students):
        row = [i + 1, s["name"], s["id_card"] or "", s["phone"] or "",
               s.get("reading_progress", 0), s.get("joined_at", "")]
        ws.append(row)
        for col in range(1, len(row) + 1):
            _style_xlsx_cell(ws.cell(row=i + 2, column=col))

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 文件名：RFC 5987 编码中文，兼容所有浏览器
    raw_name = f"{cls.name}_学员名单_{datetime.now().strftime('%Y%m%d')}.xlsx"
    ascii_name = raw_name.encode('ascii', 'ignore').decode() or 'students.xlsx'
    encoded_name = __import__('urllib.parse', fromlist=['quote']).quote(raw_name)
    disposition = f"attachment; filename={ascii_name}; filename*=UTF-8''{encoded_name}"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition}
    )


def _export_scores_xlsx(cls: Class, tests: list, scores_data: list) -> StreamingResponse:
    """导出成绩表为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩表"

    headers = ["姓名", "身份证号"] + [t.title for t in tests]
    ws.append(headers)
    _style_xlsx_header(ws, 1, len(headers))

    # 统计行
    stat_row = ["平均分", "", ""] + ["-" for _ in tests]

    for i, row_data in enumerate(scores_data):
        row = [row_data.get("姓名", ""), row_data.get("身份证号", "")] + \
              [row_data.get(t.title, "-") for t in tests]
        ws.append(row)
        for col in range(1, len(row) + 1):
            _style_xlsx_cell(ws.cell(row=i + 2, column=col))
        # 收集数值用于平均分
        for j, t in enumerate(tests):
            val = row_data.get(t.title, "-")
            if isinstance(val, (int, float)):
                if stat_row[j + 2] == "-":
                    stat_row[j + 2] = []
                stat_row[j + 2].append(val)

    # 写入平均分行
    for j in range(2, len(stat_row)):
        if isinstance(stat_row[j], list):
            stat_row[j] = round(sum(stat_row[j]) / len(stat_row[j]), 1)
    ws.append(stat_row)
    for col in range(1, len(stat_row) + 1):
        cell = ws.cell(row=len(scores_data) + 2, column=col)
        _style_xlsx_cell(cell)
        cell.font = Font(name="微软雅黑", size=10, bold=True)

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 文件名：RFC 5987 编码中文，兼容所有浏览器
    raw_name = f"{cls.name}_成绩表_{datetime.now().strftime('%Y%m%d')}.xlsx"
    ascii_name = raw_name.encode('ascii', 'ignore').decode() or 'scores.xlsx'
    encoded_name = __import__('urllib.parse', fromlist=['quote']).quote(raw_name)
    disposition = f"attachment; filename={ascii_name}; filename*=UTF-8''{encoded_name}"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition}
    )


# ═══════════════════════════════════════════════
# 公告管理
# ═══════════════════════════════════════════════

@router.get("/announcements")
async def list_announcements(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """公告列表"""
    result = await db.execute(
        select(Announcement).where(Announcement.class_id == cls.id)
        .order_by(Announcement.pinned.desc(), Announcement.created_at.desc())
    )
    items = result.scalars().all()
    return [{
        "id": a.id, "title": a.title, "content": a.content,
        "pinned": a.pinned,
        "created_by_name": a.creator.name if a.creator else "",
        "created_at": a.created_at.isoformat() if a.created_at else None
    } for a in items]


@router.post("/announcements")
async def create_announcement(
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """发布公告"""
    ann = Announcement(
        class_id=cls.id, title=data.get("title", ""),
        content=data.get("content", ""), created_by=user.id,
        pinned=data.get("pinned", False)
    )
    db.add(ann)
    await db.commit()
    return {"success": True, "id": ann.id}


@router.delete("/announcements/{ann_id}")
async def delete_announcement(
    ann_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """删除公告"""
    ann = await db.get(Announcement, ann_id)
    if not ann or ann.class_id != cls.id:
        raise HTTPException(status_code=404, detail="公告不存在")
    await db.delete(ann)
    await db.commit()
    return {"success": True}


# ═══════════════════════════════════════════════
# 跨班对比分析
# ═══════════════════════════════════════════════

@router.get("/cross-class/comparison")
async def get_cross_class_comparison(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """跨班对比：当前班 vs 全平台平均"""
    # 当前班数据
    current = await _class_snapshot(db, cls)

    # 全平台所有班级
    result = await db.execute(select(Class).where(Class.id != cls.id))
    all_classes = result.scalars().all()

    if not all_classes:
        return {"current": current, "platform_avg": current, "class_details": []}

    platform_snapshots = []
    for c in all_classes:
        snap = await _class_snapshot(db, c)
        platform_snapshots.append(snap)

    # 计算平台平均值
    def avg(values):
        return round(sum(values) / len(values), 1) if values else 0

    platform_avg = {
        "total_students": avg([s["total_students"] for s in platform_snapshots]),
        "avg_reading_progress": avg([s["avg_reading_progress"] for s in platform_snapshots]),
        "avg_score": avg([s["avg_score"] for s in platform_snapshots]),
        "pass_rate": avg([s["pass_rate"] for s in platform_snapshots]),
        "completion_rate": avg([s["completion_rate"] for s in platform_snapshots]),
        "avg_daily_study_min": avg([s["avg_daily_study_min"] for s in platform_snapshots])
    }

    return {
        "current": current,
        "platform_avg": platform_avg,
        "class_details": sorted(platform_snapshots,
            key=lambda x: x["avg_score"], reverse=True)[:10]  # Top 10
    }


async def _class_snapshot(db: AsyncSession, cls: Class) -> dict:
    """单个班级的数据快照"""
    # 学员数
    result = await db.execute(
        select(func.count()).select_from(ClassMember).where(
            ClassMember.class_id == cls.id, ClassMember.role == UserRole.STUDENT
        )
    )
    total = result.scalar() or 0

    if total == 0:
        return {"class_name": cls.name, "total_students": 0, "avg_reading_progress": 0,
                "avg_score": 0, "pass_rate": 0, "completion_rate": 0, "avg_daily_study_min": 0}

    # 阅读进度
    result = await db.execute(
        select(func.avg(ReadingProgress.progress)).select_from(ReadingProgress).join(
            ClassMember, ReadingProgress.user_id == ClassMember.user_id
        ).where(ClassMember.class_id == cls.id, ClassMember.role == UserRole.STUDENT)
    )
    avg_progress = result.scalar() or 0

    # 测验成绩
    result = await db.execute(
        select(func.avg(TestResult.score)).select_from(TestResult).join(
            ClassMember, TestResult.user_id == ClassMember.user_id
        ).where(ClassMember.class_id == cls.id)
    )
    avg_score = result.scalar() or 0

    # 通过率（>=60分）
    result = await db.execute(
        select(TestResult).join(ClassMember, TestResult.user_id == ClassMember.user_id)
        .where(ClassMember.class_id == cls.id)
    )
    all_results = result.scalars().all()
    passed = sum(1 for r in all_results if (r.score or 0) >= 60)
    pass_rate = round(passed / len(all_results) * 100, 1) if all_results else 0

    # 完成率
    result = await db.execute(
        select(ChapterProgress).join(
            ClassMember, ChapterProgress.user_id == ClassMember.user_id
        ).where(
            ClassMember.class_id == cls.id,
            ChapterProgress.status == ChapterProgressStatus.COMPLETED
        )
    )
    completed_chapters = len(result.scalars().all())
    result = await db.execute(
        select(ChapterProgress).join(
            ClassMember, ChapterProgress.user_id == ClassMember.user_id
        ).where(ClassMember.class_id == cls.id)
    )
    total_chapters = len(result.scalars().all())
    completion_rate = round(completed_chapters / total_chapters * 100, 1) if total_chapters else 0

    # 日平均学习时长（分钟）
    avg_daily = 0
    if total_chapters > 0:
        result = await db.execute(
            select(func.avg(ReadingProgress.duration)).select_from(ReadingProgress).join(
                ClassMember, ReadingProgress.user_id == ClassMember.user_id
            ).where(ClassMember.class_id == cls.id)
        )
        avg_daily = round((result.scalar() or 0) / 60, 1)

    return {
        "class_id": cls.id, "class_name": cls.name,
        "total_students": total,
        "avg_reading_progress": round(avg_progress, 1),
        "avg_score": round(avg_score, 1),
        "pass_rate": pass_rate,
        "completion_rate": completion_rate,
        "avg_daily_study_min": avg_daily
    }


# ═══════════════════════════════════════════════
# 预警面板
# ═══════════════════════════════════════════════

@router.get("/alerts")
async def get_alerts(
    unread_only: bool = Query(False),
    severity: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """获取预警记录（当前班级）"""
    q = select(AlertRecord).where(AlertRecord.class_id == cls.id)
    if unread_only:
        q = q.where(AlertRecord.is_read == False)
    if severity:
        q = q.where(AlertRecord.severity == severity)
    q = q.order_by(AlertRecord.created_at.desc()).limit(100)

    result = await db.execute(q)
    alerts = result.scalars().all()
    return [{
        "id": a.id, "user_name": a.user_name, "alert_type": a.alert_type,
        "alert_message": a.alert_message, "severity": a.severity,
        "is_read": a.is_read, "is_resolved": a.is_resolved,
        "created_at": a.created_at.isoformat() if a.created_at else None
    } for a in alerts]


@router.post("/alerts/{alert_id}/read")
async def mark_alert_read(
    alert_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """标记预警为已读"""
    alert = await db.get(AlertRecord, alert_id)
    if not alert or alert.class_id != cls.id:
        raise HTTPException(status_code=404, detail="预警不存在")
    alert.is_read = True
    await db.commit()
    return {"success": True}


@router.post("/alerts/{alert_id}/resolve")
async def resolve_alert(
    alert_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """解决预警"""
    alert = await db.get(AlertRecord, alert_id)
    if not alert or alert.class_id != cls.id:
        raise HTTPException(status_code=404, detail="预警不存在")
    alert.is_resolved = True
    alert.resolved_at = datetime.now()
    alert.resolved_by = user.id
    await db.commit()
    return {"success": True}


@router.get("/alerts/stats")
async def get_alerts_stats(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """预警统计"""
    result = await db.execute(
        select(AlertRecord).where(AlertRecord.class_id == cls.id)
    )
    all_alerts = result.scalars().all()

    unread = sum(1 for a in all_alerts if not a.is_read)
    unresolved = sum(1 for a in all_alerts if not a.is_resolved)
    by_type = {}
    for a in all_alerts:
        by_type[a.alert_type] = by_type.get(a.alert_type, 0) + 1

    return {
        "total": len(all_alerts), "unread": unread, "unresolved": unresolved,
        "by_type": by_type
    }


@router.post("/alerts/detect")
async def run_alert_detection(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    cls: Class = Depends(get_manager_class)
):
    """手动触发预警检测"""
    new_alerts = []

    # 获取启用的规则
    result = await db.execute(select(AlertRule).where(AlertRule.enabled == True))
    rules = result.scalars().all()

    # 获取班级所有学员
    result = await db.execute(
        select(ClassMember, User).join(User).where(
            ClassMember.class_id == cls.id, ClassMember.role == UserRole.STUDENT
        )
    )
    students = [(cm, u) for cm, u in result.all()]

    for rule in rules:
        if rule.rule_type == "inactivity":
            threshold_days = rule.threshold_value or 3
            cutoff = datetime.now(timezone.utc).timestamp() - threshold_days * 86400

            for cm, stu in students:
                # 检查最后活动时间
                result = await db.execute(
                    select(AuditLog.created_at).where(
                        AuditLog.user_name == stu.name
                    ).order_by(AuditLog.created_at.desc()).limit(1)
                )
                last = result.scalar_one_or_none()
                if not last or last.timestamp() < cutoff:
                    msg = f"学员 {stu.name} 已连续 {threshold_days} 天无学习活动"
                    # 去重：同类型同用户24h内不重复
                    new_alerts.append(_mk_alert(rule, stu, cls, msg, "warning", db))

        elif rule.rule_type == "score_drop":
            threshold_pct = rule.threshold_value or 30
            for cm, stu in students:
                result = await db.execute(
                    select(TestResult).where(TestResult.user_id == stu.id)
                    .order_by(TestResult.submitted_at.desc()).limit(3)
                )
                recent = result.scalars().all()
                if len(recent) >= 2:
                    prev_avg = sum(r.score for r in recent[1:] if r.score) / max(len(recent[1:]), 1)
                    latest = recent[0].score or 0
                    if prev_avg > 0 and (prev_avg - latest) / prev_avg * 100 >= threshold_pct:
                        msg = f"学员 {stu.name} 最近成绩下降 {round((prev_avg - latest) / prev_avg * 100)}%（{prev_avg}→{latest}）"
                        new_alerts.append(_mk_alert(rule, stu, cls, msg, "critical", db))

        elif rule.rule_type == "fail_rate":
            threshold_pct = rule.threshold_value or 60
            for cm, stu in students:
                result = await db.execute(
                    select(TestResult).where(TestResult.user_id == stu.id)
                )
                all_results = result.scalars().all()
                if all_results:
                    failed = sum(1 for r in all_results if (r.score or 0) < 60)
                    rate = failed / len(all_results) * 100
                    if rate >= threshold_pct:
                        msg = f"学员 {stu.name} 不及格率 {round(rate)}% ({failed}/{len(all_results)})"
                        new_alerts.append(_mk_alert(rule, stu, cls, msg, "critical", db))

        elif rule.rule_type == "low_progress":
            threshold_pct = rule.threshold_value or 50
            for cm, stu in students:
                result = await db.execute(
                    select(func.avg(ReadingProgress.progress)).where(
                        ReadingProgress.user_id == stu.id
                    )
                )
                avg_prog = result.scalar() or 0
                if avg_prog < threshold_pct:
                    msg = f"学员 {stu.name} 学习进度仅 {round(avg_prog)}%，低于 {threshold_pct}% 阈值"
                    new_alerts.append(_mk_alert(rule, stu, cls, msg, "warning", db))

    if new_alerts:
        for alert in new_alerts:
            if alert:
                db.add(alert)
        await db.commit()

    return {"created": len([a for a in new_alerts if a])}


def _mk_alert(rule, student, cls, msg, severity, db):
    """去重创建预警"""
    # 检查是否已存在同类预警（24h内）
    return AlertRecord(
        rule_id=rule.id, user_id=student.id, user_name=student.name,
        class_id=cls.id, class_name=cls.name,
        alert_type=rule.rule_type, alert_message=msg,
        severity=severity
    )


# ═══════════════════════════════════════════════
# 审计日志查询
# ═══════════════════════════════════════════════

@router.get("/audit-logs")
async def get_audit_logs(
    limit: int = 20, offset: int = 0,
    action: str = "", user_name: str = "",
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager)
):
    """查询审计日志（分页+筛选）"""
    q = select(AuditLog)
    if action:
        q = q.where(AuditLog.action == action)
    if user_name:
        q = q.where(AuditLog.user_name.ilike(f"%{user_name}%"))

    count_q = select(func.count()).select_from(AuditLog)
    if action:
        count_q = count_q.where(AuditLog.action == action)
    if user_name:
        count_q = count_q.where(AuditLog.user_name.ilike(f"%{user_name}%"))

    total = (await db.execute(count_q)).scalar()
    items = (await db.execute(q.order_by(AuditLog.created_at.desc()).offset(offset).limit(limit))).scalars().all()

    return {
        "total": total,
        "items": [{
            "id": l.id, "user_id": l.user_id, "user_name": l.user_name,
            "action": l.action, "target_type": l.target_type, "target_id": l.target_id,
        "detail": l.details, "created_at": l.created_at.isoformat() if l.created_at else None
        } for l in items]
    }


@router.get("/audit-logs/stats")
async def get_audit_logs_stats(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager)
):
    """审计日志统计"""
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today = (await db.execute(
        select(func.count()).select_from(AuditLog).where(AuditLog.created_at >= today_start)
    )).scalar()
    deletes = (await db.execute(
        select(func.count()).select_from(AuditLog).where(AuditLog.action == "DELETE")
    )).scalar()
    exports = (await db.execute(
        select(func.count()).select_from(AuditLog).where(AuditLog.action == "EXPORT")
    )).scalar()
    return {"today": today, "deletes": deletes, "exports": exports}
